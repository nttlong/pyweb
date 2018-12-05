# from openpyxl.workbook import Workbook
# from openpyxl.writer.excel import save_virtual_workbook
# from openpyxl import utils
class __excel_column__(object):
    def __init__(self):
        self.field = None
        self.caption =None
        self.datatype = None
        self.format = None
        self.index = 0
        self.lookup = None



class __excel_workbook__(object):
    def __init__(self,name):
        self.name = name
        self.columns =[__excel_column__]
        self.items = []
        self.filter = None
        self.pipeline = None
        self.source = None
        self.fileName = None

    def __feed_column__(self,other):
        from pymqr.pydocs import Fields
        col = __excel_column__()
        col.field = other[0]
        if isinstance(other[0],Fields):
            col.field =other[0].__name__
        col.caption = other[1]
        if other.__len__()>2:
            col.format = other[2]
        col.index = self.columns.__len__()+1
        self.columns.append(col)
        return self
    def __create_name_ranges(self,wb,ws):
        from openpyxl.worksheet import Worksheet
        from openpyxl.workbook import Workbook
        from openpyxl import utils
        if isinstance(wb,Workbook):
            if isinstance(ws,Worksheet):
                for col in self.columns:
                    wb.create_named_range(
                        col.field,
                        ws,
                        "$" + utils.get_column_letter(col.index) + ":$" +
                        utils.get_column_letter(col.index)
                    )
                ws.append([col.caption for col in self.columns ])
    def __lshift__(self, other):
        import pymqr.pydocs
        if self.columns.__len__()==1 and self.columns[0] == __excel_column__:
            self.columns = []
        if not isinstance(other,tuple):
            raise Exception("Column info must be tuple like this (fieldName,caption,format)")
        if type(other[0]) in [str,unicode,pymqr.pydocs.Fields]:
            self.__feed_column__(other)
        else:
            for item in other:
                self.__feed_column__(item)
        return self
    def __create_wb__(self):
        from openpyxl.workbook import Workbook
        wb = Workbook()
        ws = wb.create_sheet(self.name)
        self.__create_name_ranges(wb,ws)
        for item in self.items:
            ws.append(item)

        return wb
    def __rshift__(self, other):
        self.name = other
        return self.__create_wb__()
    def fill_data(self,data):
        import datetime
        lst = data
        x1 = datetime.datetime.now()

        if not isinstance(data,list):
            lst = list(data)
        for i in range(0,lst.__len__(),1):
            item = lst[i]
            _items = []
            for col in self.columns:
                _items.append(item.get(col.field,None))
            self.items.append(_items)
        x2 = datetime.datetime.now()

        return self
    def set_query(self,qr):
        from pymqr.pyquery import query
        from pymongo.collection import Collection
        from .JSON_PipelineParse import to_json
        if isinstance(qr,query):
            coll = qr.coll
            if isinstance(coll,Collection):
                self.source = coll.name
            self.pipeline = to_json(qr.pipeline)
            return self
        else:
            raise Exception("The param must be '{0}'".format(query))

    def set_filter(self,data):
        if not isinstance(data,dict):
            raise Exception("param must be {0}".format(dict))
        self.filter = data
        return self
    def for_django_http_export(self,download_filename = None):
        if download_filename== None:
            download_filename = self.fileName
        if download_filename == None:
            download_filename = self.source
        wb =self.__create_wb__()
        from openpyxl.writer.excel import save_virtual_workbook
        from django.http import HttpResponse
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename= "{}"'.format(download_filename + ".xlsx")
        return response
    def register(self,schema,language,app_name,view_path,file_name,username):
        """a
        Register and get download t
        :param username: icket
        :return:
        """
        if self.source == None:
            raise Exception("Can not register export without source, please call set_query")
        from .excel_exports import ExcelExportTokens,ExcelExportTokensColumnInfo
        from pymqr import query
        from django_db import getdb
        cols = []
        for col in self.columns:
            cols.append(
                (ExcelExportTokensColumnInfo<<{
                    ExcelExportTokensColumnInfo.Caption:col.caption,
                    ExcelExportTokensColumnInfo.Field:col.field,
                    ExcelExportTokensColumnInfo.Format:col.format,
                    ExcelExportTokensColumnInfo.Index:col.index,
                    ExcelExportTokensColumnInfo.Lookup: col.lookup



                })
            )
        insert_data = ExcelExportTokens <<{
            ExcelExportTokens.filter:self.filter,
            ExcelExportTokens.Language:language,
            ExcelExportTokens.AppName:app_name,
            ExcelExportTokens.Filename:file_name,
            ExcelExportTokens.Schema:schema,
            ExcelExportTokens.ViewPath:view_path,
            ExcelExportTokens.Columns:cols,
            ExcelExportTokens.Source:self.source,
            ExcelExportTokens.Pipeline:self.pipeline

        }
        qr = query(getdb(), ExcelExportTokens)
        ret, err, result = qr.insert(insert_data).commit()
        return  result.inserted_id.__str__()






def load_from_token(token):
    from pymqr import query, filters, docs
    from bson import ObjectId
    from .excel_exports import ExcelExportTokens
    from . JSON_PipelineParse import from_json
    from django_db import getdb
    qr = query(getdb(), ExcelExportTokens)
    qr.where(filters._id == ObjectId(token))

    ret_data = qr.object
    ret = excel>>ret_data.Source
    for col in ret_data.Columns:
        ret<<(col.Field,col.Caption,col.Format)
    qr_data = query(getdb(),ret_data.Source)
    qr_data.pipeline = from_json(ret_data.Pipeline)
    items = qr_data.items
    ret.fill_data(items)
    ret.fileName = ret_data.Filename

    return ret


class Excel(object):
    def __rshift__(self, other):
        return __excel_workbook__(other)
    def __floordiv__(self, other):
        x=1

excel = Excel()